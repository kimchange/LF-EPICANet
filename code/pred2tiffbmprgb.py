import torch
import h5py
import scipy.io as scio
from tifffile import imwrite
import numpy as np
import os
import utils
import openpyxl
import imageio

# datasets = ['NTIRE_Val_Real', 'NTIRE_Val_Synth']
folder = '../TestResultsx4-full-pad-ep63'
# folder = '../TestResultsx2-full-pad8ifLytro-ep60'
folder = '../TestResultsx4-20230901-patch64'
folder = '../TestResultsx2-20230831-patch64'
datasets = ['EPFL', 'HCI_new', 'HCI_old', 'INRIA_Lytro', 'Stanford_Gantry'] #os.listdir(folder)
savefolder = folder + '_tiffrgb'
# labelfolder = '../LFSSR_data/datasets'

labelfolder = '../../LFSR/LFSSR_data/datasets'

angRes = 5

savevolume = 1
calc_psnrssim = 1

class ExcelFile(object):
    def __init__(self):
        self.xlsx_file = openpyxl.Workbook()
        self.worksheet = self.xlsx_file.active
        self.worksheet.title = 'sheet1'
        self.header_list = ['Datasets', 'Scenes', 'PSNR', 'SSIM', 'LPIPS']

        self.sum = 1
        self.worksheet.cell(self.sum, 1, 'Datasets')
        self.worksheet.cell(self.sum, 2, 'Scenes')
        self.worksheet.column_dimensions['A'].width = 16
        self.worksheet.column_dimensions['B'].width = 22
        self.add_count(1)

    def write_sheet(self, test_name, LF_name, metric_name, metric_score):
        self.worksheet.cell(self.sum, 1, test_name)
        self.worksheet.cell(self.sum, 2, LF_name)

        # self.worksheet.col(self.header_list.index(metric_name)).width = 256 * 10
        self.worksheet.cell(1, self.header_list.index(metric_name)+1, metric_name)
        self.worksheet.cell(self.sum, self.header_list.index(metric_name)+1, '%.6f' % metric_score)

    def add_count(self, num):
        self.sum = self.sum + num

def LF_rgb2ycbcr(x):
    y = torch.zeros_like(x)
    y[:,0,:,:,:,:] =  65.481 * x[:,0,:,:,:,:] + 128.553 * x[:,1,:,:,:,:] +  24.966 * x[:,2,:,:,:,:] +  16.0
    y[:,1,:,:,:,:] = -37.797 * x[:,0,:,:,:,:] -  74.203 * x[:,1,:,:,:,:] + 112.000 * x[:,2,:,:,:,:] + 128.0
    y[:,2,:,:,:,:] = 112.000 * x[:,0,:,:,:,:] -  93.786 * x[:,1,:,:,:,:] -  18.214 * x[:,2,:,:,:,:] + 128.0

    y = y / 255.0
    return y


def LF_ycbcr2rgb(x):
    mat = np.array(
        [[65.481, 128.553, 24.966],
         [-37.797, -74.203, 112.0],
         [112.0, -93.786, -18.214]])
    mat_inv = np.linalg.inv(mat)
    offset = np.matmul(mat_inv, np.array([16, 128, 128]))
    mat_inv = mat_inv * 255

    y = torch.zeros_like(x)
    y[:,0,:,:,:,:] = mat_inv[0,0] * x[:,0,:,:,:,:] + mat_inv[0,1] * x[:,1,:,:,:,:] + mat_inv[0,2] * x[:,2,:,:,:,:] - offset[0]
    y[:,1,:,:,:,:] = mat_inv[1,0] * x[:,0,:,:,:,:] + mat_inv[1,1] * x[:,1,:,:,:,:] + mat_inv[1,2] * x[:,2,:,:,:,:] - offset[1]
    y[:,2,:,:,:,:] = mat_inv[2,0] * x[:,0,:,:,:,:] + mat_inv[2,1] * x[:,1,:,:,:,:] + mat_inv[2,2] * x[:,2,:,:,:,:] - offset[2]
    return y


for dataset in datasets:
    if not os.path.exists(savefolder+'/'+dataset+'/'):
        try:
            os.mkdir(savefolder+'/'+dataset+'/') 
        except:
            os.makedirs(savefolder+'/'+dataset+'/') 

psnr_testset = []
ssim_testset = []
num_testset = []

excel_file = ExcelFile()


for dataset in datasets:
    labels = os.listdir(labelfolder+'/'+dataset+'/test')
    psnr_file = []
    ssim_file = []

    if calc_psnrssim:
        txtfile = open(savefolder + '/y_psnrssim.txt', 'a')
        txtfile.write('Dataset----%10s :\n' % (dataset))
        txtfile.close()
        print('Dataset----%10s :\n' % (dataset))

    for label in labels:
        try:
            label_rgbh = h5py.File(labelfolder+'/'+dataset+'/test'+'/'+label, 'r')
            label_rgbn = np.array( label_rgbh.get('LF') ).transpose((4, 3, 2, 1, 0)) # u,v,h,w,c
        except:
            label_rgbh = scio.loadmat(labelfolder+'/'+dataset+'/test'+'/'+label)
            label_rgbn = np.array( label_rgbh.get('LF') ) # u,v,h,w,c
        (U, V, H, W, _) = label_rgbn.shape
        H = H // 4 * 4
        W = W // 4 * 4

        # Extract central angRes * angRes views
        label_rgbn = label_rgbn[(U-angRes)//2:(U+angRes)//2, (V-angRes)//2:(V+angRes)//2, 0:H, 0:W, 0:3]
        (U, V, H, W, _) = label_rgbn.shape
        label_rgbn = np.transpose(label_rgbn, (4,0,1,2,3)) # c,u,v,h,w
        # label_rgbh = h5py.File(labelfolder+'/'+dataset+'/test'+'/'+label, 'r')
        # label_rgbn = np.array( label_rgbh.get('LF') ) # c,w,h,v,u
        # label_rgbn = np.transpose(label_rgbn, (0,4,3,2,1)) # c,u,v,h,w
        label_rgbt = torch.tensor(label_rgbn).unsqueeze(0) # 1,c,u,v,h,w
        label_ycbcrt = LF_rgb2ycbcr(label_rgbt)


        pred_ycbcrt = label_ycbcrt
        pred_yh = h5py.File(folder+'/'+dataset+'/'+label[0:-4]+'.h5', 'r')
        pred_yn = np.array( pred_yh.get('LF') ) # u v h w
        pred_yt = torch.tensor(pred_yn) # u v h w
        if calc_psnrssim:
            psnr, ssim = utils.cal_metrics(label_ycbcrt[0,0,:,:,:,:], pred_yt, angRes)
            
            psnr_file.append(psnr)
            ssim_file.append(ssim)

            txtfile = open(savefolder + '/y_psnrssim.txt', 'a')
            txtfile.write('test file---%15s ,\t PSNR---%f,\t SSIM---%f\n' % (label, psnr, ssim))
            txtfile.close()
            print('test file---%15s ,\t PSNR---%f,\t SSIM---%f\n' % (label, psnr, ssim))

            excel_file.write_sheet(dataset, label[0:-4], 'PSNR', psnr)
            excel_file.write_sheet(dataset, label[0:-4], 'SSIM', ssim)
            excel_file.add_count(1)





        if savevolume: 
            pred_ycbcrt[0,0,:,:,:,:] = pred_yt

            pred_rgbt = LF_ycbcr2rgb(pred_ycbcrt).squeeze(0) # c,u,v,h,w
            pred_rgbt = pred_rgbt.permute(1,2,0,3,4 ) # u,v,c,h,w

            pred_rgbn = np.array(pred_rgbt)
            pred_rgbn = np.reshape(pred_rgbn, [pred_rgbn.shape[0]*pred_rgbn.shape[1],pred_rgbn.shape[2], pred_rgbn.shape[3], pred_rgbn.shape[4]])

            imwrite(savefolder+'/'+dataset+'/'+label[0:-4]+'.tif', pred_rgbn, imagej=True, metadata={'axes': 'ZCYX'}, compression ='zlib')

            ''' Save RGB '''
            if savefolder is not None:
                save_dir_ = savefolder + '/' +dataset+'/'+ (label[0:-4])
                if not os.path.exists(save_dir_):
                    try:
                        os.mkdir(save_dir_) 
                    except:
                        os.makedirs(save_dir_) 

                # save_dir_.mkdir(exist_ok=True)
                views_dir = save_dir_ + '/' +('views')
                if not os.path.exists(views_dir):
                    try:
                        os.mkdir(views_dir) 
                    except:
                        os.makedirs(views_dir) 

                # views_dir.mkdir(exist_ok=True)

                # save the center view
                LF_out = LF_ycbcr2rgb(pred_ycbcrt)
                LF_out = (LF_out.squeeze(0).permute(1, 2, 3, 4, 0).cpu().detach().numpy().clip(0, 1) * 255).astype('uint8')
                path = str(save_dir_) + '/' + label[0:-4] + '_SAI.bmp'
                img = LF_out[angRes//2, angRes//2, :, :, :]
                imageio.imwrite(path, img)


                # save all views
                for i in range(angRes):
                    for j in range(angRes):
                        path = str(views_dir) + '/' + label[0:-4] + '_' + str(i) + '_' + str(j) + '.bmp'
                        img = LF_out[i, j, :, :, :]
                        imageio.imwrite(path, img)
                    pass
        

    
    if calc_psnrssim:
        psnr_testset.append(float(np.array(psnr_file).mean()))
        ssim_testset.append(float(np.array(ssim_file).mean()))
        num_testset.append(len(psnr_file))
        txtfile = open(savefolder + '/y_psnrssim.txt', 'a')
        txtfile.write('Dataset----%10s,\t test Number---%d ,\t PSNR---%f,\t SSIM---%f\n' % (dataset, num_testset[-1], psnr_testset[-1], ssim_testset[-1]))
        txtfile.close()
        print('Dataset----%10s,\t test Number---%d ,\t PSNR---%f,\t SSIM---%f\n' % (dataset, num_testset[-1], psnr_testset[-1], ssim_testset[-1]))

        excel_file.write_sheet(dataset, 'Average', 'PSNR', psnr_testset[-1])
        excel_file.write_sheet(dataset, 'Average', 'SSIM', ssim_testset[-1])
        excel_file.add_count(2)
        

if calc_psnrssim:
    psnr_avg = sum([psnr_testset[ii]*num_testset[ii] for ii in range(len(num_testset))]) / sum(num_testset)
    ssim_avg = sum([ssim_testset[ii]*num_testset[ii] for ii in range(len(num_testset))]) / sum(num_testset)
    txtfile = open(savefolder + '/y_psnrssim.txt', 'a')
    txtfile.write('Total testset,\t test Number---%d ,\t PSNR---%f,\t SSIM---%f\n' % (sum(num_testset), psnr_avg, ssim_avg))
    txtfile.close()
    print('Total testset,\t test Number---%d ,\t PSNR---%f,\t SSIM---%f\n' % (sum(num_testset), psnr_avg, ssim_avg))

    excel_file.write_sheet('ALL', 'Average', 'PSNR', psnr_avg)
    excel_file.write_sheet('ALL', 'Average', 'SSIM', ssim_avg)
    print('The mean psnr on testsets is %.5f, mean ssim is %.5f' % (psnr_avg, ssim_avg))
    excel_file.xlsx_file.save(str(savefolder) + '/evaluation.xlsx')
        # lfh = h5py.File(folder+'/'+dataset+'/'+file, 'r')
        # lfn = np.array( lfh.get('LF') ) # c,w,h,v,u


        # lfn = np.transpose(lfn, (4,3,0,2,1)) # u,v,c,h,w
        # print(f'{dataset}/{file} lfn.shape is {lfn.shape}')
        # lfn = np.reshape(lfn, [lfn.shape[0]*lfn.shape[1],lfn.shape[2], lfn.shape[3], lfn.shape[4]])

        # imwrite(folder+'_tif/'+dataset+'/'+file[0:-4]+'.tif', lfn, imagej=True, metadata={'axes': 'ZCYX'}, compression ='zlib')

